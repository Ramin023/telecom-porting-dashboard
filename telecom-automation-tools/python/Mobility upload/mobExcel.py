from openpyxl import load_workbook, Workbook
from mobExcelMethd import mobExcelMethd
import os

class mobExcel:
    
    def __init__(self, name, savePath, filePath):
        self.name = name
        self.savePath = savePath
        
        current_directory = os.getcwd()
        self.filePath = filePath
         
        
    def addService(self):
        wb_read = load_workbook(filename=self.filePath)
        mo = wb_read['Sheet1'].iter_rows(min_row=2, values_only=True)
        
        wb_write = Workbook()
        sc = wb_write.active

        sc_headers = ["ID", "Macnum", "Order #", "Marketplace Order ID", "In-House Account", "User Name", "Bundle Type",
              "Access Type", "Provider", "Product", "Service", "Contract Term", "SIM Card Status", "SIM Carrier",
              "SIM Card/ICCID", "TN/MDN", "Throttling", "Throttling Type", "IP Type", "IP Address", "Subnet Mask",
              "Default Gateway", "IP Passthrough", "Group Plan ID", "Data Plan Type", "Data Plan", "Talk", "Text",
              "Roaming", "Travel Plan", "Data Rate Plan", "Voice Rate Plan", "Text Rate Plan", "Equipment ID",
              "Shipping Date", "Macnum Address", "Business Name", "Contact Name", "Contact Phone #", "Email Address",
              "Service Address1", "Service Address2", "City", "State \\ Province", "ZIP", "Country", "Shipping Notes",
              "User Added", "SF OPP ID", "Upload Speed", "Download Speed", "Mobile Network", "Traffic Management",
              "APN", "Carrier Account Identifier"]

        mobExcelMethd.font_header(sc, sc_headers)

        n = 2
        for row in mo:
            sc['A' + str(n)] = n-1
            sc['B' + str(n)] = str(row[2]).strip()
            sc['C' + str(n)] = str(row[3]).strip()
            sc['E' + str(n)] = "No"
            sc['F' + str(n)] = row[6].strip()
            sc['G' + str(n)] = "None"
            sc['H' + str(n)] = "Cellular"
            sc['J' + str(n)] = "Voice"
            sc['L' + str(n)] = "2 Yr"
            sc['M' + str(n)] = "Activated"
            sc['O' + str(n)] = str(row[9]).strip()
            sc['P' + str(n)] = str(row[13]).strip()
            sc['Q' + str(n)] = "None"
            sc['S' + str(n)] = "Dynamic"
            sc['Y' + str(n)] = "Standard"
            sc['AA' + str(n)] = "Unlimited"
            sc['AB' + str(n)] = "Unlimited SMS"
            sc['AE' + str(n)] = "GENDT"
            sc['AF' + str(n)] = "GENVC"
            sc['AG' + str(n)] = "GENTX"
            sc['AI' + str(n)] = mobExcelMethd.convertDateToStr(row[0])
            sc['AJ' + str(n)] = "Yes"
            sc['AW' + str(n)] = str(row[12]).strip()
            
            carrier = row[10].replace(" ","").upper().split('>')
            
            if carrier[1] == "VZ":
                sc['I' + str(n)] = "Ready Wireless"
                sc['N' + str(n)] = "Ready Wireless"
                sc['AC' + str(n)] = "Disabled"
                sc['AD' + str(n)] = "Disabled"
            elif carrier[1] == "ATT":
                sc['I' + str(n)] = "APEX"
                sc['N' + str(n)] = "ATT"
                sc['AC' + str(n)] = "Measured Rate Plan"
                sc['AD' + str(n)] = "Day Pass"
             
            #reset a plan to correct format, like as "3 GB"
            if row[11].strip()[0].isdigit():
                plan = row[11].strip().replace(" ", "")[:-2] + ' ' + row[11].strip()[-2:].upper() 
            else:
                plan = row[11].strip()
                
            if plan[0].isdigit():
                sc['K' + str(n)] = "Pooled Connection"
                sc['Z' + str(n)] = plan

            else:
                sc['K' + str(n)] = "Individual connection"
                sc['Z' + str(n)] = "Unlimited"
                
            n +=1

        save_path = mobExcelMethd.saveEx(self.savePath, "Service", self.name.replace(" ", ""))
        wb_write.save(save_path)
        wb_write.close()
        wb_read.close()
        print(f"Mobility Service Done - {save_path}")
    
 
    def addCharge(self):
        wb_read = load_workbook(filename=self.filePath)
        mo = wb_read['Sheet1'].iter_rows(min_row=2, values_only=True)
        
        wb_write = Workbook()
        ch = wb_write.active

        ch_headers = ["ID", "Macnum", "Charge Type", "Utilities ID", "Add Date", "Invoice Number", "End Date", 
                      "Frequency", "Prorate", "Overwrite Description", "Overwrite Amount Flag", "Overwrite Amount", 
                      "Mobility ID", "Mobility BTN", "SIM ID", "User Added"]

        mobExcelMethd.font_header(ch, ch_headers)

        n = 2
        for row in mo:
            ch['A' + str(n)] = n-1
            ch['B' + str(n)] = str(row[2]).strip()
            ch['C' + str(n)] = "S"
            ch['E' + str(n)] = mobExcelMethd.convertDateToStr(row[0])
            ch['H' + str(n)] = "Monthly"
            ch['I' + str(n)] = "Yes"
            ch['J' + str(n)] = str(row[9]).strip()
            ch['K' + str(n)] = "No"
            ch['M' + str(n)] = str(row[13]).strip()
            ch['O' + str(n)] = str(row[9]).strip()
            
            
            carrier = row[10].replace(" ","").upper().split('>')
            
            #reset a plan to correct format, like as "3 GB"
            if row[11].strip()[0].isdigit():
                plan = row[11].strip().replace(" ", "")[:-2] + ' ' + row[11].strip()[-2:].upper() 
            else:
                plan = row[11].strip()
            
            if carrier[1] == 'ATT':
                if plan == '100 MB':
                    ch['D' + str(n)] ='148626044'
                elif plan == '1 GB':
                    ch['D' + str(n)] = '148626048'
                elif plan == '2 GB':
                    ch['D' + str(n)] = '148626052'
                elif plan == '3 GB':
                    ch['D' + str(n)] = '148626056'
                elif plan == '4 GB':
                    ch['D' + str(n)] = '148626060'
                elif plan == '5 GB':
                    ch['D' + str(n)] = '148626064'
                elif plan == '6 GB':
                    ch['D' + str(n)] = '148626068'
                elif plan == '10 GB':
                    ch['D' + str(n)] = '149004797'
                elif 'Unlimited' in plan and 'Elite' in plan:
                    ch['D' + str(n)] = '149090111'
                elif 'Unlimited' in plan in plan:
                    ch['D' + str(n)] = '149090110'
                else:
                    print(f"{carrier[1]} {plan} is not valid plan for charge - ID {n-1}")
            elif carrier[1] == 'VZ':
                if plan == '100 MB':
                    ch['D' + str(n)] = '148626034'
                elif plan == '1 GB':
                    ch['D' + str(n)] = '148626119'
                elif plan == '2 GB':
                    ch['D' + str(n)] = '148626121'
                elif plan == '4 GB':
                    ch['D' + str(n)] = '148626123'
                elif carrier[0] == 'VZ':
                    if plan == '3 GB':
                        ch['D' + str(n)] = '149394427'
                    elif plan == '5 GB':
                        ch['D' + str(n)] = '149394429'
                    elif 'Unlimited' in plan and 'Premium' in plan:
                        ch['D' + str(n)] = '149394433' 
                    elif 'Unlimited' in plan in plan:
                        ch['D' + str(n)] = '149394434'
                    else:
                        print(f"{carrier[1]} {plan} is not valid plan for charge - ID {n-1}")
                else:
                    if plan == '3 GB':
                        ch['D' + str(n)] = '149394430'
                    elif plan == '5 GB':
                        ch['D' + str(n)] = '149394431'
                    elif 'Unlimited' in plan and 'Premium' in plan:
                        ch['D' + str(n)] = '149394433' 
                    elif 'Unlimited' in plan in plan:
                        ch['D' + str(n)] = '149394434'
                    else:
                        print(f"{carrier[1]} {plan} is not valid plan for charge - ID {n-1}")
            else:
                print(f"{carrier[1]} {plan} is not valid plan for charge - ID {n-1}")
            n +=1

        save_path = mobExcelMethd.saveEx(self.savePath, "Charge", self.name.replace(" ", ""))
        wb_write.save(save_path)
        wb_write.close()
        wb_read.close()
        print(f"Mobility Charge Done - {save_path}")           


    def addEquipment(self):
        wb_read = load_workbook(filename=self.filePath)
        mo = wb_read['Sheet1'].iter_rows(min_row=2, values_only=True)

        wb_write = Workbook()
        eq = wb_write.active

        eq_headers = ["ID", "Macnum", "Order #", "Device Status", "Product Family", "Equipment Policy Group",
                      "BYOD Subtype", "BYOD Description", "Return Date", "Bundle Type", "Product", "Manufacturer",
                      "Description", "Model #", "Device Username", "Equip ID Type", "Equip ID Value", "Managed",
                      "Node Name", "Back-up Device", "Granite Part", "Custom Configuration",
                      "Other Configuration / Remote Management",
                      "Add-On Service Type", "Add on Product", "Provider", "Description", "Activation Date",
                      "Shipping Date",
                      "Macnum Address", "Business Name", "Contact Name", "Contact Phone #", "Email Address",
                      "Service Address1",
                      "Service Address2", "City", "State \ Province", "ZIP", "Country", "Shipping Notes",
                      "Service Link",
                      "Service Link 2", "Service Link 3", "Service Link 4", "Service Link 5", "Equipment ID",
                      "User Added"]

        mobExcelMethd.font_header(eq, eq_headers)

        n = 2
        for row in mo:
            eq['A' + str(n)] = n - 1
            eq['B' + str(n)] = str(row[2]).strip()
            eq['C' + str(n)] = str(row[3]).strip()
            eq['D' + str(n)] = "Activated"
            eq['E' + str(n)] = "Wireless Mobility"
            eq['F' + str(n)] = "BYOD"
            eq['G' + str(n)] = "Supported"
            eq['H' + str(n)] = row[7].strip()
            eq['J' + str(n)] = "None"
            eq['K' + str(n)] = "Smart Phone"
            eq['L' + str(n)] = "Other"
            eq['M' + str(n)] = "Other"
            eq['N' + str(n)] = "Other"
            eq['P' + str(n)] = "IMEI"
            eq['Q' + str(n)] = str(row[8]).strip()
            eq['R' + str(n)] = "No"
            eq['V' + str(n)] = "Other"
            eq['W' + str(n)] = "Other"
            eq['AC' + str(n)] = mobExcelMethd.convertDateToStr(row[0])
            eq['AD' + str(n)] = "Yes"
            eq['AP' + str(n)] = str(row[9]).strip()
            n += 1

        save_path = mobExcelMethd.saveEx(self.savePath, "Equipment", self.name.replace(" ", ""))
        wb_write.save(save_path)
        wb_write.close()
        wb_read.close()
        print(f"Mobility Equipment Done - {save_path}")
        
        

